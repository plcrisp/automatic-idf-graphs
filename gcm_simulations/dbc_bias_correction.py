"""
Este módulo implementa o método **DBC (Distribution-Based Correction)**, uma técnica estatística 
para **correção de viés em dados climáticos simulados por modelos GCMs (Global Climate Models)**. 
Ele é especialmente útil na adequação de projeções futuras com base em séries históricas observacionais e simuladas.

## O que é correção de viés?

Modelos climáticos frequentemente apresentam **viés sistemático**, ou seja, tendências consistentes 
de desvio em relação aos valores reais observados — como superestimar chuvas intensas ou subestimar 
temperaturas mínimas. A **correção de viés (bias correction)** busca reduzir esses desvios usando 
métodos estatísticos baseados em dados históricos confiáveis.

## O que é o método DBC?

O **DBC (Distribution-Based Correction)** é um método de correção de viés baseado na comparação das 
distribuições estatísticas dos dados observacionais e simulados. Seu objetivo é ajustar os valores 
simulados do modelo para que sua distribuição se aproxime da distribuição dos dados observados.

A abordagem utilizada neste módulo consiste em:

1. **Extrair pares de dados históricos** (observacional e simulado) para um período comum.
2. Determinar um **limiar (threshold)** com base nos percentis dos dados simulados.
3. Identificar os valores simulados acima desse limiar e calcular **fatores de correção por percentil**, 
    com base na relação entre os dados observacionais e simulados.
4. Aplicar esses fatores de correção aos dados futuros simulados, ajustando-os conforme a curva de distribuição histórica observada.

Esse método é especialmente útil para cenários onde é necessário manter a coerência estatística 
entre os dados simulados e os observados, como em estudos de impacto climático e modelagens hidrológicas.
"""

from openpyxl import Workbook, load_workbook
import xlrd
from pathlib import Path
from typing import List, Tuple
import scipy.stats as st
import pandas as pd

"""
--------------------------------------------------------------------------------------------------------------
-------------------------------- CORREÇÃO DE VIÉS BASEADA NA DISTRIBUIÇÃO ------------------------------------
--------------------------------------------------------------------------------------------------------------
"""


def load_gcm_data(file_path: str) -> pd.DataFrame:
    """
    Carrega dados GCM de um arquivo CSV e remove valores nulos.
    
    Args:
        file_path (str): Caminho do arquivo CSV.
        
    Returns:
        pd.DataFrame: Dados carregados e limpos.
    """
    data_df = pd.read_csv(file_path)
    return data_df.dropna()



def save_to_excel(data_df: pd.DataFrame, file_path: str, header: bool = False) -> None:
    """
    Salva DataFrame em um arquivo Excel.
    
    Args:
        data_df (pd.DataFrame): DataFrame a ser salvo.
        file_path (str): Caminho do arquivo de saída.
        header (bool): Se deve incluir cabeçalho ou não.
    """
    data_df.to_excel(file_path, index=False, header=header)



def get_observations_and_model_values(
    input_sheet,
    obs_col: int = 0,
    year_col: int = 1,
    nyears: int = 20
) -> Tuple[List[float], List[float]]:
    """
    Extrai dados observacionais e simulados do modelo baseado no número de anos.
    
    Args:
        input_sheet: Planilha aberta com os dados.
        obs_col (int): Índice da coluna com os dados observacionais.
        year_col (int): Índice da coluna com os anos.
        nyears (int): Número de anos para considerar no histórico.
        
    Returns:
        Tuple[List[float], List[float]]: Lista dos dados observacionais e simulados.
    """
    obs_values = []
    sim_values = []
    y = 0
    x1 = 0
    year = 0

    while y < nyears + 1:
        precip_value = input_sheet.cell_value(x1, obs_col)
        if precip_value > 0.001:
            obs_values.append(precip_value)
        x1 += 1
        new_year = input_sheet.cell_value(x1, year_col)
        if new_year > year or x1 >= input_sheet.nrows - 1:
            y += 1
            year = new_year
        if x1 >= input_sheet.nrows:
            break

    for x in range(0, x1):
        precip_value = input_sheet.cell_value(x, obs_col + 1)
        sim_values.append(precip_value)

    sim_values.sort()
    obs_values.sort()

    threshold = sim_values[len(sim_values) - len(obs_values)]
    sim_above_threshold = [val for val in sim_values if val > threshold]

    return obs_values, sim_above_threshold, threshold



def compute_percentile_factors(
    obs_values: List[float],
    sim_above_threshold: List[float],
    num_percentiles: int = 1000
) -> List[float]:
    """
    Calcula fatores de correção por percentil entre dados observacionais e simulados.
    
    Args:
        obs_values (List[float]): Valores observados.
        sim_above_threshold (List[float]): Valores simulados acima do threshold.
        num_percentiles (int): Número de percentis a serem usados.
        
    Returns:
        List[float]: Fatores de correção por percentil.
    """
    percentile_factors = []
    for p in range(num_percentiles):
        idx_obs = int((p + 1) / num_percentiles * (len(obs_values) - 1))
        idx_sim = int((p + 1) / num_percentiles * (len(sim_above_threshold) - 1))
        percentile_factors.append(obs_values[idx_obs] / sim_above_threshold[idx_sim])
    return percentile_factors



def correct_historical_data(
    input_sheet,
    data_col: int,
    percentile_factors: List[float],
    threshold: float,
    sim_above_threshold: List[float],
    output_sheet,
    start_row: int = 0
):
    """
    Aplica a correção aos dados históricos com base nos fatores calculados.
    
    Args:
        input_sheet: Planilha original com dados.
        data_col (int): Coluna atual sendo processada.
        percentile_factors (List[float]): Fatores de correção por percentil.
        threshold (float): Threshold usado.
        sim_above_threshold (List[float]): Valores simulados acima do threshold.
        output_sheet: Planilha onde salvar os resultados corrigidos.
        start_row (int): Linha inicial para começar a escrever os dados.
    """
    for x in range(start_row, input_sheet.nrows):
        precip_value = input_sheet.cell_value(x, data_col)
        percentile_level = 0

        if precip_value > threshold:
            while (
                precip_value > sim_above_threshold[int((percentile_level + 1) / len(percentile_factors) * (len(sim_above_threshold) - 1))] and
                percentile_level < len(percentile_factors) - 1
            ):
                percentile_level += 1
            corrected_value = percentile_factors[percentile_level] * precip_value
        else:
            corrected_value = 0

        output_sheet.write(x - start_row, data_col - 4, corrected_value)
        output_sheet.write(x - start_row, data_col - 3, input_sheet.cell_value(x, 1))  # ano
        output_sheet.write(x - start_row, data_col - 2, input_sheet.cell_value(x, 2))  # mês
        output_sheet.write(x - start_row, data_col - 1, input_sheet.cell_value(x, 3))  # dia



def dbc_calib_valid(name_gcm: str, input_dir: str = 'GCM_data/dbc_bias_correction', nyears: int = 20) -> None:
    """
    Calibra o modelo usando DBC Bias Correction com base em dados históricos.
    
    Args:
        name_gcm (str): Nome do modelo GCM.
        input_dir (str): Diretório contendo os arquivos de entrada.
        nyears (int): Número de anos para calibração.
    """
    input_file = Path(input_dir) / f'{name_gcm}_baseline_to_dbc.csv'
    data_df = load_gcm_data(str(input_file))
    
    excel_file = Path(input_dir) / f'{name_gcm}_baseline_to_dbc.xls'
    save_to_excel(data_df, str(excel_file), header=False)

    input_book = xlrd.open_workbook(str(excel_file))
    input_sheet = input_book.sheet_by_index(0)

    wb_factors = Workbook()
    wb_corrected = Workbook()
    wb_thresholds = Workbook()

    sheet_factors = wb_factors.active
    sheet_corrected = wb_corrected.active
    sheet_thresholds = wb_thresholds.active

    NUM_PERCENTILES = 1000
    num_cols = input_sheet.ncols

    for data_col in range(4, num_cols):
        print(f'Number of GCMs corrected: {data_col - 3} out of {num_cols - 4}')
        obs_values, sim_above_threshold, threshold = get_observations_and_model_values(
            input_sheet, obs_col=0, year_col=1, nyears=nyears
        )
        percentile_factors = compute_percentile_factors(obs_values, sim_above_threshold, NUM_PERCENTILES)

        for i, factor in enumerate(percentile_factors):
            sheet_factors.cell(row=i + 1, column=data_col - 3, value=factor)
        sheet_thresholds.cell(row=1, column=data_col - 3, value=threshold)

        correct_historical_data(input_sheet, data_col, percentile_factors, threshold, sim_above_threshold, sheet_corrected)

    wb_factors.save(Path(input_dir) / f'{name_gcm}_percentfactors.xls')
    wb_corrected.save(Path(input_dir) / f'{name_gcm}_baseline_validation.xls')
    wb_thresholds.save(Path(input_dir) / f'{name_gcm}_tresholds.xls')



def correct_future_by_dbc(name_gcm, scenario):
    """
    Aplica a correção de viés (bias correction) DBC aos dados futuros do modelo GCM,
    utilizando os fatores percentílicos e thresholds calculados durante a calibração histórica.

    A função lê os dados futuros do modelo, aplica a correção baseada em percentis dos dados simulados
    e salva os resultados em um novo arquivo Excel contendo os valores corrigidos junto com as datas
    (ano, mês, dia).

    Args:
        name_gcm (str): Nome do modelo GCM (ex: 'HadGEM2-ES', 'MIROC5').
        scenario (str): Cenário climático futuro (ex: 'rcp45', 'rcp85').

    Returns:
        None: A função salva o resultado em um arquivo Excel, não retorna valor.

    Exemplo:
        correct_future_by_dbc('HadGEM2-ES', 'rcp45')
        Arquivo salvo em: GCM_data/dbc_bias_correction/HadGEM2-ES_rcp45_DBC_daily_simple.xls
    """
    path_future = f'GCM_data/dbc_bias_correction/{name_gcm}_{scenario}_to_dbc.xlsx'
    path_thresholds = f'GCM_data/dbc_bias_correction/{name_gcm}_tresholds.xls'
    path_factors = f'GCM_data/dbc_bias_correction/{name_gcm}_percentfactors.xls'

    op_future = xlrd.open_workbook(path_future)
    tresh_book = xlrd.open_workbook(path_thresholds)
    perc_book = xlrd.open_workbook(path_factors)

    all45 = op_future.sheet_by_index(0)
    thold = tresh_book.sheet_by_index(0)
    per = perc_book.sheet_by_index(0)

    output_wb = Workbook()
    sheet_corrected = output_wb.active
    sheet_corrected.title = "Corrected Data"

    for data_col in range(3, all45.ncols):
        print(f'Coluna {data_col - 3} de {all45.ncols - 4}')
        sim_values = []
        year_counter = 0
        year = all45.cell_value(0, 0)
        year0 = year

        for row in range(all45.nrows):
            value = all45.cell_value(row, data_col)
            if value > 0:
                sim_values.append(value)

        sim_values.sort()
        sim_above_threshold = sim_values[:]

        cont = 0
        for row in range(all45.nrows):
            precip_value = all45.cell_value(row, data_col) if all45.cell_value(row, data_col) > 0 else 0
            percentile_level = 0

            threshold = thold.cell_value(0, data_col - 3)
            percentile_factors = [per.cell_value(p, data_col - 3) for p in range(per.nrows)]

            if precip_value > threshold:
                while (
                    precip_value > sim_above_threshold[int((percentile_level + 1) / per.nrows * (len(sim_above_threshold) - 1))] and
                    percentile_level < per.nrows - 1
                ):
                    percentile_level += 1
                corrected_value = percentile_factors[percentile_level] * precip_value
            else:
                corrected_value = 0

            sheet_corrected.cell(row=cont + 1, column=data_col - 3, value=corrected_value)
            sheet_corrected.cell(row=cont + 1, column=data_col - 2, value=all45.cell_value(row, 0))  # ano
            sheet_corrected.cell(row=cont + 1, column=data_col - 1, value=all45.cell_value(row, 2))  # mês
            sheet_corrected.cell(row=cont + 1, column=data_col, value=all45.cell_value(row, 3))      # dia

            cont += 1

    output_path = f'GCM_data/dbc_bias_correction/{name_gcm}_{scenario}_DBC_daily_simple.xls'
    output_wb.save(output_path)
    print(f"Arquivo salvo em: {output_path}")

    
    
    